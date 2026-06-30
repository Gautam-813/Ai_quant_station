"""
Master Report Generator
=======================
Merges all existing daily_report_*.xlsx files in daily_reports/ into one
master Excel file covering the full period.

Usage:
    # From backend/ directory:
    python -m scripts.generate_master_report

    # Dry-run (no email):
    python -m scripts.generate_master_report --dry-run
"""
import argparse
import os
import re
import sys
from datetime import datetime
from pathlib import Path

# Ensure backend/ is on sys.path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

REPORTS_DIR = Path(__file__).resolve().parent.parent / "daily_reports"


# ── Parsing ────────────────────────────────────────────────────────────────

def _parse_daily_xlsx(filepath: Path) -> dict | None:
    """Parse a daily_report_YYYY-MM-DD.xlsx file. Returns {summary, trades} or None."""
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [WARN] Cannot open {filepath.name}: {e}")
        return None

    if "Daily Report" not in wb.sheetnames:
        print(f"  [WARN] {filepath.name}: no 'Daily Report' sheet, skipping")
        wb.close()
        return None

    ws = wb["Daily Report"]
    rows = list(ws.iter_rows(values_only=True))

    # Extract date from row 2 (e.g. "Monday, June 11, 2026   |   Report Generated: 23:51 UTC")
    date_str = ""
    if len(rows) > 1 and rows[1][0]:
        full = str(rows[1][0])
        date_part = full.split("|")[0].strip()
        try:
            dt = datetime.strptime(date_part, "%A, %B %d, %Y")
            date_str = dt.strftime("%Y-%m-%d")
        except ValueError:
            date_str = date_part

    # Summary rows 4-10 (0-indexed: 3-9)
    # Row 5 (idx 4): Total Trades
    # Row 6 (idx 5): Wins
    # Row 7 (idx 6): Losses
    # Row 8 (idx 7): Win Rate
    # Row 9 (idx 8): Daily P&L
    # Row 10 (idx 9): Best Prompt
    summary = {"date": date_str}
    summary_labels = ["total_trades", "wins", "losses", "win_rate", "pnl", "best_prompt"]
    for i, key in enumerate(summary_labels):
        idx = 4 + i
        if idx < len(rows):
            val = rows[idx][3] if len(rows[idx]) > 3 else None  # Column D (index 3)
            if key == "win_rate" and val is not None:
                val = str(val).replace("%", "").strip()
                try:
                    val = float(val)
                except ValueError:
                    val = 0.0
            elif key == "pnl" and val is not None:
                val = str(val).replace("$", "").replace(",", "").strip()
                try:
                    val = float(val)
                except ValueError:
                    val = 0.0
            elif key in ("total_trades", "wins", "losses") and val is not None:
                try:
                    val = int(val)
                except (ValueError, TypeError):
                    val = 0
            summary[key] = val

    # Trades: start after "Today's Trades" header + header row
    # Header at row 13 (idx 12), column headers at row 14 (idx 13), data starts row 15 (idx 14)
    trades = []
    trade_start = 14  # 0-indexed
    for r in rows[trade_start:]:
        if not r or not r[0]:
            continue
        ticket = r[0]
        if ticket is None or str(ticket).strip() == "":
            continue
        if str(ticket).strip().startswith("Prompt"):
            break  # reached "Prompt Performance" section
        trade = {
            "ticket": str(ticket) if ticket else "",
            "symbol": str(r[1]) if len(r) > 1 and r[1] else "",
            "direction": str(r[2]) if len(r) > 2 and r[2] else "",
            "entry_price": float(r[3]) if len(r) > 3 and r[3] is not None else None,
            "exit_price": float(r[4]) if len(r) > 4 and r[4] is not None else None,
            "lot_size": float(r[5]) if len(r) > 5 and r[5] is not None else 0.0,
            "profit": float(r[6]) if len(r) > 6 and r[6] is not None else 0.0,
            "prompt": str(r[7]) if len(r) > 7 and r[7] else "",
            "result": str(r[8]) if len(r) > 8 and r[8] else "",
        }
        trade["date"] = date_str
        trades.append(trade)

    wb.close()
    return {"summary": summary, "trades": trades, "date": date_str}


# ── Aggregation ────────────────────────────────────────────────────────────

def _compute_overall(trades: list[dict]) -> dict:
    closed = [t for t in trades if t.get("result", "") and t["result"] != "Open"]
    wins = sum(1 for t in closed if t["result"] == "Win")
    losses = sum(1 for t in closed if t["result"] == "Loss")
    pnl = sum(t["profit"] for t in closed)
    total = len(closed)
    win_rate = round(wins / total * 100, 1) if total > 0 else 0.0
    avg_profit = round(pnl / total, 2) if total > 0 else 0.0
    return {
        "total_trades": total,
        "wins": wins,
        "losses": losses,
        "win_rate": win_rate,
        "pnl": round(pnl, 2),
        "avg_profit": avg_profit,
    }


def _compute_prompt_stats(trades: list[dict]) -> list[dict]:
    closed = [t for t in trades if t.get("result", "") and t["result"] != "Open"]
    groups: dict[str, dict] = {}
    for t in closed:
        key = t.get("prompt", "?") or "?"
        if key not in groups:
            groups[key] = {"prompt": key, "wins": 0, "losses": 0, "pnl": 0.0}
        groups[key]["pnl"] += t["profit"]
        if t["result"] == "Win":
            groups[key]["wins"] += 1
        else:
            groups[key]["losses"] += 1

    stats = []
    for g in groups.values():
        total = g["wins"] + g["losses"]
        win_pct = round(g["wins"] / total * 100, 1) if total > 0 else 0.0
        avg_p = round(g["pnl"] / total, 2) if total > 0 else 0.0
        stats.append({
            "prompt": g["prompt"],
            "wins": g["wins"],
            "losses": g["losses"],
            "wl": f"{g['wins']}W / {g['losses']}L",
            "pnl": round(g["pnl"], 2),
            "win_pct": win_pct,
            "avg_profit": avg_p,
        })
    stats.sort(key=lambda x: x["pnl"], reverse=True)
    return stats


def _compute_daily_breakdown(all_days: list[dict]) -> list[dict]:
    """Build day-by-day breakdown from parsed file summaries."""
    result = []
    for day in all_days:
        d = {
            "date": day.get("date", ""),
            "trades": day.get("summary", {}).get("total_trades", 0),
            "wins": day.get("summary", {}).get("wins", 0),
            "losses": day.get("summary", {}).get("losses", 0),
            "pnl": day.get("summary", {}).get("pnl", 0.0),
        }
        total = d["wins"] + d["losses"]
        d["win_rate"] = round(d["wins"] / total * 100, 1) if total > 0 else 0.0
        try:
            dt = datetime.strptime(d["date"], "%Y-%m-%d")
            d["day_name"] = dt.strftime("%A")
        except (ValueError, TypeError):
            d["day_name"] = ""
        result.append(d)
    return result


# ── Excel Generation ──────────────────────────────────────────────────────

def _generate_master_excel(overall: dict, daily_breakdown: list[dict],
                           all_trades: list[dict], prompt_stats: list[dict]) -> str:
    wb = Workbook()

    thin = Side(style="thin")
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12, color="1F4E79")
    blue_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # ── Sheet 1: Master Summary ──
    ws = wb.active
    ws.title = "Master Summary"

    for col, w in zip("ABCDEFGHIJ", [14, 14, 10, 12, 12, 10, 12, 14, 14, 10]):
        ws.column_dimensions[col].width = w

    r = 1
    ws.merge_cells("A1:J1")
    ws["A1"] = "AI Quant Station — Full Period Master Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    r = 2
    ws.merge_cells(f"A{r}:J{r}")
    date_range = ""
    if daily_breakdown:
        start = daily_breakdown[0]["date"]
        end = daily_breakdown[-1]["date"]
        date_range = f"{start} — {end}"
    ws[f"A{r}"] = f"Period: {date_range}   |   Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws[f"A{r}"].font = Font(size=10, italic=True)
    ws[f"A{r}"].alignment = Alignment(horizontal="center")

    r = 4
    ws[f"A{r}"] = "Overall Performance"
    ws[f"A{r}"].font = section_font

    labels = [
        ("Trading Days", len(daily_breakdown)),
        ("Total Trades (Closed)", overall["total_trades"]),
        ("Wins", overall["wins"]),
        ("Losses", overall["losses"]),
        ("Win Rate", f"{overall['win_rate']}%"),
        ("Total P&L", f"${overall['pnl']:.2f}"),
        ("Avg Profit / Trade", f"${overall['avg_profit']:.2f}"),
    ]
    for i, (label, value) in enumerate(labels):
        r = 5 + i
        ws[f"B{r}"] = label
        ws[f"B{r}"].font = header_font
        ws[f"D{r}"] = value
        ws[f"B{r}"].border = Border(bottom=thin)
        ws[f"D{r}"].border = Border(bottom=thin)

    # ── Sheet 2: Day-by-Day ──
    ws2 = wb.create_sheet("Day-by-Day")
    for col, w in zip("ABCDEFGHIJ", [14, 14, 10, 10, 10, 10, 14]):
        ws2.column_dimensions[col].width = w

    r2 = 1
    ws2.merge_cells("A1:G1")
    ws2["A1"] = "Daily Breakdown"
    ws2["A1"].font = section_font

    r2 = 3
    headers = ["Date", "Day", "Trades", "Wins", "Losses", "Win %", "P&L"]
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=r2, column=c, value=h)
        cell.font = header_font
        cell.fill = blue_fill
        cell.border = Border(bottom=thin)

    prev_row = r2
    for dd in daily_breakdown:
        r2 += 1
        ws2.cell(row=r2, column=1, value=dd["date"])
        ws2.cell(row=r2, column=2, value=dd.get("day_name", ""))
        ws2.cell(row=r2, column=3, value=dd["trades"])
        ws2.cell(row=r2, column=4, value=dd["wins"])
        ws2.cell(row=r2, column=5, value=dd["losses"])
        ws2.cell(row=r2, column=6, value=f"{dd['win_rate']}%")
        pnl_cell = ws2.cell(row=r2, column=7, value=dd["pnl"])
        pnl_cell.number_format = "+$#,##0.00;-$#,##0.00"
        for c in range(1, 8):
            ws2.cell(row=r2, column=c).border = Border(bottom=thin)

    # Bar chart
    if daily_breakdown:
        chart_start = prev_row + 1  # first data row
        chart_end = chart_start + len(daily_breakdown) - 1

        chart = BarChart()
        chart.type = "col"
        chart.title = "Daily P&L — Full Period"
        chart.y_axis.title = "P&L ($)"
        chart.x_axis.title = "Day"
        chart.style = 10

        data_ref = Reference(ws2, min_col=7, min_row=chart_start, max_row=chart_end)
        cats_ref = Reference(ws2, min_col=2, min_row=chart_start, max_row=chart_end)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        chart.width = 22
        chart.height = 14

        series = chart.series[0]
        series.graphicalProperties.solidFill = "2563EB"

        ws2.add_chart(chart, f"A{r2 + 2}")

    # ── Sheet 3: All Trades ──
    ws3 = wb.create_sheet("All Trades")
    for col, w in zip("ABCDEFGHIJK", [10, 14, 10, 10, 12, 12, 8, 12, 10, 12, 10]):
        ws3.column_dimensions[col].width = w

    r3 = 1
    ws3.merge_cells("A1:K1")
    ws3["A1"] = "All Trades — Full Period"
    ws3["A1"].font = section_font

    r3 = 3
    headers3 = ["Ticket", "Date", "Symbol", "Direction", "Entry", "Exit", "Lot", "Profit", "Prompt#", "Result", "Day"]
    for c, h in enumerate(headers3, 1):
        cell = ws3.cell(row=r3, column=c, value=h)
        cell.font = header_font
        cell.fill = blue_fill
        cell.border = Border(bottom=thin)

    for t in all_trades:
        r3 += 1
        ws3.cell(row=r3, column=1, value=t.get("ticket", ""))
        ws3.cell(row=r3, column=2, value=t.get("date", ""))
        ws3.cell(row=r3, column=3, value=t.get("symbol", ""))
        ws3.cell(row=r3, column=4, value=t.get("direction", ""))
        ws3.cell(row=r3, column=5, value=t.get("entry_price"))
        ws3.cell(row=r3, column=6, value=t.get("exit_price"))
        ws3.cell(row=r3, column=7, value=t.get("lot_size"))
        profit_cell = ws3.cell(row=r3, column=8, value=t.get("profit", 0.0))
        profit_cell.number_format = "+$#,##0.00;-$#,##0.00"
        ws3.cell(row=r3, column=9, value=t.get("prompt", ""))
        ws3.cell(row=r3, column=10, value=t.get("result", ""))
        try:
            dt = datetime.strptime(t.get("date", ""), "%Y-%m-%d")
            ws3.cell(row=r3, column=11, value=dt.strftime("%A"))
        except (ValueError, TypeError):
            ws3.cell(row=r3, column=11, value="")
        for c in range(1, 12):
            ws3.cell(row=r3, column=c).border = Border(bottom=thin)

    # ── Sheet 4: Prompt Performance ──
    ws4 = wb.create_sheet("Prompt Performance")
    for col, w in zip("ABCDEFGH", [14, 14, 10, 10, 14, 12, 14]):
        ws4.column_dimensions[col].width = w

    r4 = 1
    ws4.merge_cells("A1:G1")
    ws4["A1"] = "Prompt Performance — Full Period"
    ws4["A1"].font = section_font

    r4 = 3
    prompt_headers = ["Prompt", "Wins", "Losses", "W/L", "P&L", "Win %", "Avg Profit"]
    for c, h in enumerate(prompt_headers, 1):
        cell = ws4.cell(row=r4, column=c, value=h)
        cell.font = header_font
        cell.fill = blue_fill
        cell.border = Border(bottom=thin)

    for ps in prompt_stats:
        r4 += 1
        ws4.cell(row=r4, column=1, value=ps["prompt"])
        ws4.cell(row=r4, column=2, value=ps["wins"])
        ws4.cell(row=r4, column=3, value=ps["losses"])
        ws4.cell(row=r4, column=4, value=ps["wl"])
        ws4.cell(row=r4, column=5, value=ps["pnl"])
        ws4.cell(row=r4, column=6, value=f"{ps['win_pct']}%")
        ws4.cell(row=r4, column=7, value=ps["avg_profit"])
        for c in range(1, 8):
            ws4.cell(row=r4, column=c).border = Border(bottom=thin)

    filename = f"master_report_{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
    filepath = str(REPORTS_DIR / filename)
    wb.save(filepath)
    print(f"  [OK] Master report saved: {filepath}")
    return filepath


# ── Email ──────────────────────────────────────────────────────────────────

def _send_email(filepath: str, overall: dict) -> bool:
    """Send the master report via SMTP."""
    try:
        from app.core.config import settings
    except ImportError:
        print("  [WARN] Cannot import settings — skipping email")
        return False

    sender = settings.REPORT_EMAIL
    raw = settings.REPORT_RECIPIENT_EMAIL
    recipients = [r.strip() for r in raw.split(",") if r.strip()] if raw else []
    smtp_server = settings.SMTP_SERVER
    smtp_port = settings.SMTP_PORT
    password = settings.REPORT_EMAIL_PASSWORD

    if not all([sender, recipients, smtp_server, smtp_port, password]):
        print("  [WARN] Email not fully configured — skipping")
        return False

    import smtplib
    from email.mime.application import MIMEApplication
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    subject = f"Master Trade Report — Full Period ({datetime.utcnow().strftime('%Y-%m-%d')})"
    body = (
        f"AI Quant Station — Full Period Master Report\n"
        f"{datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC\n\n"
        f"Overall Performance:\n"
        f"  Total Trades (Closed): {overall['total_trades']}\n"
        f"  Wins: {overall['wins']}  /  Losses: {overall['losses']}\n"
        f"  Win Rate: {overall['win_rate']}%\n"
        f"  Total P&L: ${overall['pnl']:.2f}\n\n"
        f"Report attached.\n— AI Quant Station"
    )

    try:
        msg = MIMEMultipart()
        msg["Subject"] = subject
        msg["From"] = sender
        msg["To"] = ", ".join(recipients)
        msg.attach(MIMEText(body, "plain"))

        with open(filepath, "rb") as f:
            att = MIMEApplication(f.read(), _subtype="xlsx")
            att.add_header("Content-Disposition", "attachment", filename=os.path.basename(filepath))
            msg.attach(att)

        if smtp_port == 465:
            with smtplib.SMTP_SSL(smtp_server, smtp_port, timeout=30) as server:
                server.login(sender, password)
                server.send_message(msg)
        else:
            with smtplib.SMTP(smtp_server, smtp_port, timeout=30) as server:
                server.starttls()
                server.login(sender, password)
                server.send_message(msg)

        print(f"  [OK] Email sent to {', '.join(recipients)}")
        return True
    except Exception as e:
        print(f"  [WARN] Email failed: {e}")
        return False


# ── Main ───────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate master trade report from daily .xlsx files")
    parser.add_argument("--dry-run", action="store_true", help="Generate report but skip email")
    args = parser.parse_args()

    if not REPORTS_DIR.is_dir():
        print(f"[ERR] Reports directory not found: {REPORTS_DIR}")
        sys.exit(1)

    # Find all daily report files
    files = sorted(REPORTS_DIR.glob("daily_report_*.xlsx"))
    if not files:
        print(f"[ERR] No daily_report_*.xlsx files found in {REPORTS_DIR}")
        sys.exit(1)

    print(f"Found {len(files)} daily report files in {REPORTS_DIR}")
    print()

    # Parse all files
    all_days = []
    all_trades = []
    for fp in files:
        print(f"  Reading: {fp.name}...")
        parsed = _parse_daily_xlsx(fp)
        if parsed:
            all_days.append(parsed)
            all_trades.extend(parsed["trades"])
            print(f"    → {len(parsed['trades'])} trades, P&L ${parsed['summary'].get('pnl', 0):.2f}")
        else:
            print(f"    → skipped")

    if not all_trades:
        print("\n[ERR] No trades found in any file — aborting")
        sys.exit(1)

    print(f"\n  Total: {len(all_days)} days, {len(all_trades)} trades")

    # Compute aggregate stats
    overall = _compute_overall(all_trades)
    daily_breakdown = _compute_daily_breakdown(all_days)
    prompt_stats = _compute_prompt_stats(all_trades)

    print(f"  Overall P&L: ${overall['pnl']:.2f} ({overall['wins']}W / {overall['losses']}L, {overall['win_rate']}%)")

    # Generate Excel
    filepath = _generate_master_excel(overall, daily_breakdown, all_trades, prompt_stats)

    # Email (unless dry-run)
    if args.dry_run:
        print("\n  [SKIP] --dry-run: email not sent")
    else:
        print()
        _send_email(filepath, overall)

    print("\nDone.")


if __name__ == "__main__":
    main()
