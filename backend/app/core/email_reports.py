"""
Daily Trade Report — Excel + Email
====================================
Flow:
  1. Try MT5 connector for today's [AUTOPILOT] trades (primary, real trades)
  2. Retry every 5min if MT5 offline (max 3 attempts)
  3. Fallback to local AutopilotTrade DB if MT5 still offline
  4. Generate .xlsx matching daily_report_demo.xlsx format
  5. Email to configured recipient
  6. Archive .xlsx to daily_reports/ folder
"""

import asyncio
import json
import logging
import os
import re
import smtplib
from datetime import datetime, timezone, timedelta
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import httpx
from apscheduler.schedulers.asyncio import AsyncIOScheduler

from ..core.config import settings

logger = logging.getLogger(__name__)

REPORTS_DIR = Path(__file__).resolve().parent.parent.parent / "daily_reports"
REPORTS_DIR.mkdir(exist_ok=True)

scheduler = AsyncIOScheduler()


# ── Helpers ────────────────────────────────────────────────────────────────

def _now() -> datetime:
    return datetime.now(timezone.utc)


def _today_start() -> datetime:
    return _now().replace(hour=0, minute=0, second=0, microsecond=0)


def _today_str() -> str:
    return _today_start().strftime("%Y-%m-%d")


def _day_name() -> str:
    return _now().strftime("%A, %B %d, %Y")


# ── Step 1: Fetch trades from MT5 connector ───────────────────────────────

async def _fetch_mt5_trades() -> list[dict] | None:
    """Fetch today's [AUTOPILOT] trades from MT5 connector.
    Returns list of trade dicts, or None if MT5 unreachable."""
    mt5_url = settings.MT5_CONNECTOR_URL
    if not mt5_url:
        return None

    mt5_base = mt5_url.rstrip("/")
    headers = {}
    if settings.MT5_API_TOKEN:
        headers["Authorization"] = f"Bearer {settings.MT5_API_TOKEN}"

    try:
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.get(
                f"{mt5_base}/history",
                params={"hours": 48},
                headers=headers,
                timeout=15,
            )
        if resp.status_code != 200:
            return None

        mt5_deals = resp.json().get("deals", [])
    except Exception:
        return None

    # Filter [AUTOPILOT] trades
    autopilot_pids = {
        d.get("position_id")
        for d in mt5_deals
        if d.get("position_id") and (d.get("comment") or "").strip().startswith("[AUTOPILOT]")
    }
    auto_deals = [d for d in mt5_deals if d.get("position_id") in autopilot_pids]

    pos_map: dict = {}
    for deal in auto_deals:
        pid = deal.get("position_id")
        if not pid:
            continue
        if pid not in pos_map:
            pos_map[pid] = {"open": None, "close": None}
        if deal.get("entry") == "OPEN":
            pos_map[pid]["open"] = deal
        else:
            pos_map[pid]["close"] = deal

    today = _today_str()
    trades = []
    for pid, pair in pos_map.items():
        open_deal = pair["open"]
        close_deal = pair["close"]
        if not open_deal:
            continue

        deal_time = open_deal.get("time", "")
        try:
            deal_dt = datetime.strptime(deal_time, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
        except (ValueError, TypeError):
            continue
        if deal_dt < _today_start():
            continue

        direction = open_deal.get("direction", "")
        profit = close_deal.get("profit", 0) if close_deal else 0
        comment = open_deal.get("comment", "") or ""
        prompt_match = re.search(r"P(\d+)", comment)
        prompt_number = int(prompt_match.group(1)) if prompt_match else None

        close_comment = (close_deal.get("comment", "") or "").lower() if close_deal else ""
        if close_deal:
            if "sl" in close_comment:
                res_type = "Loss"
            elif "tp" in close_comment:
                res_type = "Win"
            elif profit > 0:
                res_type = "Win"
            else:
                res_type = "Loss"
        else:
            res_type = "Open"

        trades.append({
            "ticket": pid,
            "symbol": open_deal.get("symbol", ""),
            "direction": direction,
            "entry_price": open_deal.get("price"),
            "exit_price": close_deal.get("price") if close_deal else None,
            "lot_size": open_deal.get("volume", 0),
            "profit": round(float(profit) if profit else 0.0, 2),
            "prompt_number": prompt_number,
            "result": res_type,
            "executed_at": open_deal.get("time", ""),
        })

    trades.sort(key=lambda t: t.get("executed_at", ""))
    return trades


# ── Step 1b: Fallback — fetch from local DB ───────────────────────────────

async def _fetch_local_trades() -> list[dict]:
    """Fallback: read today's autopilot trades from local DB."""
    from sqlalchemy import select
    from ..core.database import AsyncSessionLocal
    from ..models.ai_memory import AutopilotTrade

    trades = []
    try:
        async with AsyncSessionLocal() as session:
            result = await session.execute(
                select(AutopilotTrade).where(
                    AutopilotTrade.executed_at >= _today_start()
                ).order_by(AutopilotTrade.executed_at)
            )
            for row in result.scalars().all():
                profit = row.profit or 0.0
                if row.result in ("TP_HIT",):
                    res_type = "Win"
                elif row.result in ("SL_HIT",):
                    res_type = "Loss"
                elif profit > 0:
                    res_type = "Win"
                else:
                    res_type = "Loss"

                trades.append({
                    "ticket": row.mt5_ticket or row.id,
                    "symbol": row.symbol,
                    "direction": row.direction,
                    "entry_price": row.entry_price,
                    "exit_price": row.exit_price,
                    "lot_size": row.lot_size,
                    "profit": round(profit, 2),
                    "prompt_number": row.prompt_number,
                    "result": res_type,
                    "executed_at": row.executed_at.strftime("%Y-%m-%d %H:%M:%S") if row.executed_at else "",
                })
    except Exception as e:
        logger.error(f"[Report] Local DB fallback failed: {e}")

    return trades


# ── Step 2: Compute summary stats ─────────────────────────────────────────

def _compute_summary(trades: list[dict]) -> dict:
    closed = [t for t in trades if t["result"] != "Open"]
    wins = sum(1 for t in closed if t["result"] == "Win")
    losses = sum(1 for t in closed if t["result"] == "Loss")
    pnl = sum(t["profit"] for t in closed)
    total = len(closed)

    win_rate = round(wins / total * 100, 1) if total > 0 else 0.0

    # Best prompt by P&L
    prompt_pnl: dict[str, float] = {}
    prompt_counts: dict[str, dict] = {}
    for t in closed:
        pn = t.get("prompt_number")
        key = f"#{pn}" if pn and pn > 0 else f"Custom-{abs(pn)}" if pn else "?"
        prompt_pnl[key] = prompt_pnl.get(key, 0) + t["profit"]
        if key not in prompt_counts:
            prompt_counts[key] = {"w": 0, "l": 0}
        if t["result"] == "Win":
            prompt_counts[key]["w"] += 1
        else:
            prompt_counts[key]["l"] += 1

    best_prompt = ""
    if prompt_pnl:
        best_key = max(prompt_pnl, key=prompt_pnl.get)
        best_w = prompt_counts[best_key]["w"]
        best_l = prompt_counts[best_key]["l"]
        best_prompt = f"{best_key}  (+${prompt_pnl[best_key]:.2f}, {best_key}W / {best_key}L)"

    return {
        "total_trades": total,
        "wins": wins,
        "losses": losses,
        "win_rate": win_rate,
        "pnl": round(pnl, 2),
        "best_prompt": best_prompt,
    }


def _compute_prompt_stats(trades: list[dict]) -> list[dict]:
    closed = [t for t in trades if t["result"] != "Open"]
    groups: dict[str, dict] = {}
    for t in closed:
        pn = t.get("prompt_number")
        key = f"#{pn}" if pn and pn > 0 else f"Custom-{abs(pn)}" if pn else "?"
        if key not in groups:
            groups[key] = {"prompt": key, "wins": 0, "losses": 0, "pnl": 0.0}
        groups[key]["pnl"] += t["profit"]
        if t["result"] == "Win":
            groups[key]["wins"] += 1
        else:
            groups[key]["losses"] += 1

    stats = []
    for g in groups.values():
        total = g["wins"] + g["losses"]
        win_pct = round(g["wins"] / total * 100, 1) if total > 0 else 0.0
        avg_profit = round(g["pnl"] / total, 2) if total > 0 else 0.0
        stats.append({
            "prompt": g["prompt"],
            "wins": g["wins"],
            "losses": g["losses"],
            "wl": f"{g['wins']}W / {g['losses']}L",
            "pnl": round(g["pnl"], 2),
            "win_pct": win_pct,
            "avg_profit": avg_profit,
        })
    stats.sort(key=lambda x: x["pnl"], reverse=True)
    return stats


# ── Step 3: Generate Excel (.xlsx) ────────────────────────────────────────

def _generate_excel(summary: dict, trades: list[dict], prompt_stats: list[dict]) -> str:
    """Generate .xlsx matching daily_report_demo.xlsx format. Returns filepath."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    # Column widths
    for col, w in zip("ABCDEFGHIJ", [12, 12, 10, 12, 12, 10, 12, 14, 14, 10]):
        ws.column_dimensions[col].width = w

    thin = Side(style="thin")
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12, color="1F4E79")
    blue_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    r = 1
    # ── Title ──
    ws.merge_cells("A1:J1")
    ws["A1"] = "AI Quant Station — Daily Trade Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    # ── Date line ──
    r = 2
    ws.merge_cells(f"A{r}:J{r}")
    ws[f"A{r}"] = f"{_day_name()}   |   Report Generated: {_now().strftime('%H:%M UTC')}"
    ws[f"A{r}"].font = Font(size=10, italic=True)
    ws[f"A{r}"].alignment = Alignment(horizontal="center")

    # ── Today's Performance ──
    r = 4
    ws[f"A{r}"] = "Today's Performance"
    ws[f"A{r}"].font = section_font

    labels = [
        ("Total Trades", summary["total_trades"]),
        ("Wins", summary["wins"]),
        ("Losses", summary["losses"]),
        ("Win Rate", f"{summary['win_rate']}%"),
        ("Daily P&L", f"${summary['pnl']:.2f}"),
        ("Best Prompt", summary["best_prompt"]),
    ]
    for i, (label, value) in enumerate(labels):
        r = 5 + i
        ws[f"B{r}"] = label
        ws[f"B{r}"].font = header_font
        ws[f"D{r}"] = value
        ws[f"B{r}"].border = Border(bottom=thin)
        ws[f"D{r}"].border = Border(bottom=thin)

    # ── Today's Trades ──
    r = 13
    ws[f"A{r}"] = "Today's Trades"
    ws[f"A{r}"].font = section_font

    r += 1
    headers = ["Ticket", "Symbol", "Direction", "Entry", "Exit", "Lot", "Profit", "Prompt#", "Result"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = header_font
        cell.fill = blue_fill
        cell.border = Border(bottom=thin)

    for trade in trades:
        r += 1
        ws.cell(row=r, column=1, value=trade["ticket"])
        ws.cell(row=r, column=2, value=trade["symbol"])
        ws.cell(row=r, column=3, value=trade["direction"])
        ws.cell(row=r, column=4, value=trade["entry_price"])
        ws.cell(row=r, column=5, value=trade["exit_price"])
        ws.cell(row=r, column=6, value=trade["lot_size"])
        profit_cell = ws.cell(row=r, column=7, value=trade["profit"])
        profit_cell.number_format = "+$#,##0.00;-$#,##0.00"
        pn = trade.get("prompt_number")
        ws.cell(row=r, column=8, value=f"#{pn}" if pn else "")
        ws.cell(row=r, column=9, value=trade["result"])
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = Border(bottom=thin)

    # ── Prompt Performance ──
    r += 2
    ws[f"A{r}"] = "Prompt Performance"
    ws[f"A{r}"].font = section_font

    r += 1
    prompt_headers = ["Prompt", "W/L", "P&L", "Win %", "Avg Profit"]
    for c, h in enumerate(prompt_headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = header_font
        cell.fill = blue_fill
        cell.border = Border(bottom=thin)

    for ps in prompt_stats:
        r += 1
        ws.cell(row=r, column=1, value=ps["prompt"])
        ws.cell(row=r, column=2, value=ps["wl"])
        ws.cell(row=r, column=3, value=ps["pnl"])
        ws.cell(row=r, column=4, value=f"{ps['win_pct']}%")
        ws.cell(row=r, column=5, value=ps["avg_profit"])
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = Border(bottom=thin)

    # ── Footer ──
    r += 2
    ws.merge_cells(f"A{r}:J{r}")
    ws[f"A{r}"] = "AI Quant Station — Automated Trading Report | Confidential"
    ws[f"A{r}"].font = Font(size=9, italic=True, color="808080")
    ws[f"A{r}"].alignment = Alignment(horizontal="center")

    # Save
    filename = f"daily_report_{_today_str()}.xlsx"
    filepath = str(REPORTS_DIR / filename)
    wb.save(filepath)
    logger.info(f"[Report] Excel saved: {filepath}")
    return filepath


# ── Step 4: Email ─────────────────────────────────────────────────────────

async def _send_email(filepath: str, summary: dict) -> bool:
    """Send the .xlsx report via SendGrid HTTPS API or SMTP fallback. Returns True on success."""
    sender = settings.REPORT_EMAIL
    raw_recipients = settings.REPORT_RECIPIENT_EMAIL
    recipients = [r.strip() for r in raw_recipients.split(",") if r.strip()] if raw_recipients else []

    if not all([sender, recipients]):
        logger.warning("[Report] Email not configured — skipping")
        return False

    # Try SendGrid first (HTTPS API, port 443 — works everywhere)
    sg_key = os.environ.get("SENDGRID_API_KEY") or ""
    if sg_key:
        try:
            import base64
            async with httpx.AsyncClient(timeout=30) as client:
                with open(filepath, "rb") as f:
                    file_b64 = base64.b64encode(f.read()).decode()

                for email in recipients:
                    payload = {
                        "personalizations": [{"to": [{"email": email}], "subject": f"Daily Trade Report — {_today_str()}"}],
                        "from": {"email": sender},
                        "content": [{"type": "text/plain", "value": (
                            f"AI Quant Station — Daily Trade Report\n"
                            f"{_day_name()}\n\n"
                            f"Today's Performance:\n"
                            f"  Total Trades: {summary['total_trades']}\n"
                            f"  Wins: {summary['wins']}  /  Losses: {summary['losses']}\n"
                            f"  Win Rate: {summary['win_rate']}%\n"
                            f"  P&L: ${summary['pnl']:.2f}\n\n"
                            f"Report attached.\n— AI Quant Station"
                        )}],
                        "attachments": [{
                            "content": file_b64,
                            "filename": os.path.basename(filepath),
                            "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            "disposition": "attachment",
                        }],
                    }
                    resp = await client.post(
                        "https://api.sendgrid.com/v3/mail/send",
                        json=payload,
                        headers={
                            "Authorization": f"Bearer {sg_key}",
                            "Content-Type": "application/json",
                        },
                    )
                    if resp.status_code not in (200, 201, 202):
                        logger.error(f"[Report] SendGrid failed for {email}: {resp.status_code} {resp.text[:200]}")
                        return False

            logger.info(f"[Report] Email sent via SendGrid to {', '.join(recipients)}")
            return True
        except Exception as e:
            logger.error(f"[Report] SendGrid failed: {e}, falling back to SMTP...")

    # Fallback: SMTP (direct port 587/465)
    smtp_server = settings.SMTP_SERVER
    smtp_port = settings.SMTP_PORT
    password = settings.REPORT_EMAIL_PASSWORD

    if not all([smtp_server, smtp_port, password]):
        logger.warning("[Report] SMTP not configured either — email skipped")
        return False

    try:
        msg = MIMEMultipart()
        msg["Subject"] = f"Daily Trade Report — {_today_str()}"
        msg["From"] = sender
        msg["To"] = ", ".join(recipients)

        body = (
            f"AI Quant Station — Daily Trade Report\n"
            f"{_day_name()}\n\n"
            f"Today's Performance:\n"
            f"  Total Trades: {summary['total_trades']}\n"
            f"  Wins: {summary['wins']}  /  Losses: {summary['losses']}\n"
            f"  Win Rate: {summary['win_rate']}%\n"
            f"  P&L: ${summary['pnl']:.2f}\n\n"
            f"Report attached as: {os.path.basename(filepath)}\n"
            f"— AI Quant Station"
        )
        msg.attach(MIMEText(body, "plain"))

        with open(filepath, "rb") as f:
            attachment = MIMEApplication(f.read(), _subtype="xlsx")
            attachment.add_header(
                "Content-Disposition", "attachment",
                filename=os.path.basename(filepath),
            )
            msg.attach(attachment)

        if smtp_port == 465:
            with smtplib.SMTP_SSL(smtp_server, smtp_port, timeout=30) as server:
                server.login(sender, password)
                server.send_message(msg)
        else:
            with smtplib.SMTP(smtp_server, smtp_port, timeout=30) as server:
                server.starttls()
                server.login(sender, password)
                server.send_message(msg)

        logger.info(f"[Report] Email sent via SMTP to {', '.join(recipients)}")
        return True
    except Exception as e:
        logger.error(f"[Report] Email failed: {e}")
        return False


# ── Orchestrator ──────────────────────────────────────────────────────────

async def run_daily_report():
    """Main entry point: MT5 → retry → fallback → Excel → email → archive."""
    logger.info("[Report] === Running Daily Report ===")

    # Step 1: Try MT5 connector (primary)
    trades = None
    source = "MT5 connector"

    for attempt in range(3):
        trades = await _fetch_mt5_trades()
        if trades is not None:
            break
        if attempt < 2:
            wait = (attempt + 1) * 300  # 5min, 10min
            logger.warning(f"[Report] MT5 offline (attempt {attempt+1}/3), retrying in {wait//60}min...")
            await asyncio.sleep(wait)

    # Step 1b: Fallback to local DB
    if trades is None:
        logger.warning("[Report] MT5 unreachable — falling back to local DB")
        trades = await _fetch_local_trades()
        source = "local DB"

    if not trades:
        logger.warning("[Report] No trades found today — generating empty report")
        source = "no data"

    logger.info(f"[Report] Source: {source} | {len(trades)} trades found")

    # Step 2: Compute stats
    summary = _compute_summary(trades)
    prompt_stats = _compute_prompt_stats(trades)

    # Step 3: Generate Excel
    filepath = _generate_excel(summary, trades, prompt_stats)

    # Step 4: Email
    await _send_email(filepath, summary)


# ── Scheduler start / stop ────────────────────────────────────────────────

def start_report_scheduler():
    """Register daily report job at 23:50 UTC."""
    if not scheduler.running:
        scheduler.add_job(
            run_daily_report,
            trigger="cron",
            hour=23,
            minute=50,
            timezone="UTC",
            id="daily_report",
            replace_existing=True,
        )
        scheduler.start()
        logger.info("[Report] Daily report scheduler started (23:50 UTC)")


def shutdown_report_scheduler():
    if scheduler.running:
        scheduler.shutdown(wait=False)
        logger.info("[Report] Scheduler shut down")
